"""XLSX extractor — openpyxl 3.1.5 (LOCKED-06 approved, pure-Python).

Returns ``(text, list[AttachmentRef])``. Iterates worksheets and dumps
each row as TAB-separated cell values (a deterministic, structure-
preserving form that LLMs read well). Embedded image extraction in
openpyxl 3.1.5 is best-effort: openpyxl exposes ``ws._images`` with
``Image`` instances whose ``ref`` attribute is sometimes a BytesIO and
sometimes a path — we handle the BytesIO case and skip the rest.
XLSX files rarely embed meaningful images for vision routing, so this
trade-off is fine.
"""
from __future__ import annotations

import io
from pathlib import Path
from typing import Tuple

from nyrahost.attachments import AttachmentRef
from nyrahost.extractors._common import (
    assert_ooxml_within_zip_budget,
    ingest_image_bytes,
)


def extract_xlsx(
    src: Path, *, project_saved: Path
) -> Tuple[str, list[AttachmentRef]]:
    assert_ooxml_within_zip_budget(src)

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as e:
        raise ValueError(
            "openpyxl not installed — cannot extract XLSX documents"
        ) from e

    try:
        # data_only=True returns formula results rather than the
        # `=SUM(A1:A10)` literal — what the LLM cares about is the
        # current value in the cell.
        wb = load_workbook(filename=str(src), data_only=True, read_only=False)
    except InvalidFileException as e:
        raise ValueError(f"malformed XLSX {src.name}: {e}") from e
    except (OSError, ValueError) as e:
        raise ValueError(f"unreadable XLSX {src.name}: {e}") from e
    except Exception as e:  # zipfile.BadZipFile and similar
        raise ValueError(f"unreadable XLSX {src.name}: {e}") from e

    text_chunks: list[str] = []
    image_refs: list[AttachmentRef] = []

    for ws in wb.worksheets:
        # Sheet header line — gives the LLM context when multi-sheet.
        text_chunks.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = [
                "" if v is None else str(v)
                for v in row
            ]
            # Skip rows that are entirely empty so a sparse sheet
            # doesn't pad the prompt with thousands of blank lines.
            if any(c for c in cells):
                text_chunks.append("\t".join(cells))
        # Best-effort image scrape.
        for img in getattr(ws, "_images", []):
            ref_attr = getattr(img, "ref", None)
            blob: bytes | None = None
            if isinstance(ref_attr, io.BytesIO):
                blob = ref_attr.getvalue()
            else:
                # path-on-disk or anchor-only Image instance — skip.
                continue
            if blob:
                emitted = ingest_image_bytes(
                    blob, project_saved=project_saved
                )
                if emitted is not None:
                    image_refs.append(emitted)

    wb.close()
    text = "\n".join(text_chunks)
    return text, image_refs
